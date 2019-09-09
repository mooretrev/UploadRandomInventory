from openpyxl.styles import Border, Font, Side, PatternFill
from openpyxl.styles.colors import BLACK, YELLOW
from  openpyxl import load_workbook
import Config as cg
from tkinter import messagebox

class ExportExcel:
	def __init__(self, filename, df, col_name_material_no, col_name_material_desc, col_name_num_parts,
	             col_name_total_value, col_name_storage_location, col_name_plant, col_name_batch,
	             col_name_currency, col_name_total_price,
	             col_name_manufacturer_no = 'null'):
		#load template
		try:
			self.wb = load_workbook(filename=cg.export_template)
		except Exception as exc:
			message = 'Critical Error: Report template could not be loaded. This exe will have to be recompiled. The .spec needs to include template location before complication. \nException Details\n {}'.format(str(exc))
			messagebox.showerror('Critial Error', message)
			exit()
		sheet_names = self.wb.get_sheet_names()
		name = sheet_names[0]
		self.ws = self.wb[name]

		#declaration
		self.filename = filename
		self.df = df
		col_name_material_cost_avg = 'Cost Avg'
		self.column_letter_to_column_names = {col_name_plant: 'A', col_name_batch: 'B',
		                                      col_name_material_desc: 'C', col_name_material_no: 'D',
		                                      col_name_num_parts: 'E',
		                                      col_name_material_cost_avg: 'F', col_name_currency: 'G',
		                                      col_name_total_price: 'H'}

		#actions methods
		self.add_cost_avg(col_name_material_cost_avg, col_name_num_parts, col_name_total_price)
		self._convert_pandas_to_excel()
		self._insert_row_formulas('J', 'E', '-', 'I')
		self._insert_row_formulas('K', 'J', '*', 'F')
		self._insert_summation_formulas('E')
		self._insert_summation_formulas('I')
		self._insert_summation_formulas('H')
		self._insert_summation_formulas('J')
		self._insert_summation_formulas('K')
		self._border_column('I')
		self._border_column('L')
		self._yellow_fill('I')
		self._format_money_numbers('F')
		self._format_money_numbers('H')
		self._format_money_numbers('K')
		self._export_excel()

	def _convert_pandas_to_excel(self):
		for dic_key in self.column_letter_to_column_names.keys():
			index = 2
			for value in self.df[dic_key]:
				cell_location = self.column_letter_to_column_names[dic_key] + str(index)
				self.ws[cell_location] = value
				self._format_cell(cell_location)

				index += 1

	def _insert_row_formulas(self, insert_column, column1, operator, column2):
		for i in range(self.df.shape[0]):
			cell_location = insert_column + str(i + 2)
			formula = '=' + column1 + str(i + 2)  + operator  + column2 + str(i + 2)
			self.ws[cell_location] = formula
			self._format_cell(cell_location)

	def _insert_summation_formulas(self, insert_column):
		formula = '=SUM(' + insert_column + '2:' + insert_column + str(self.df.shape[0] + 1) + ')'
		cell_location = insert_column + str(self.df.shape[0] + 2)
		self.ws[cell_location] = formula
		self._format_bottom_row(cell_location)

	def _format_bottom_row(self, cell_location):
		self.ws[cell_location].font = Font(name='calibri', size=8, bold=True)

	def _format_cell(self, cell_location):
		border_side = Side(style='thin', color=BLACK)
		self.ws[cell_location].border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
		self.ws[cell_location].font = Font(name='calibri', size=8)

	def _yellow_fill(self, column):
		for i in range(2, self.df.shape[0]+2):
			cell_location = column + str(i)
			self.ws[cell_location].fill = PatternFill(fgColor=YELLOW, fill_type='solid')

	def _border_column(self, column):
		for i in range(2, self.df.shape[0]+2):
			cell_location = column + str(i)
			self._format_cell(cell_location)

	def _format_money_numbers(self, column):
		for i in range(2, self.df.shape[0] + 3):
			cell_location = column + str(i)
			format = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'
			self.ws[cell_location].number_format = format

	def add_cost_avg(self, col_name_cost, col_name_count, col_name_value):
		self.df[col_name_cost] = self.df[col_name_value] / self.df[col_name_count]

	def _export_excel(self):
		try:
			self.wb.save(self.filename)
		except Exception as exc:
			print('Here')
			message = 'Report could not be saved. This could be due to an open excel. Close all excels are rerun the report. Attempted file save location: {}\nException Report\n{}'.format(self.filename, str(exc))
			messagebox.showerror('Error in Saving', message)
			exit()